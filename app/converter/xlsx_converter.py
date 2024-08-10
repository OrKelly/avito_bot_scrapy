import datetime
import io
import os.path
import re
import uuid
from typing import Any

from openpyxl.styles import NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.converter.styles import get_style
from app.settings.core import config


class ImportXlsx:

    # Данные для стиля заголовков таблицы
    style_config_header = {
        'name': 'header',
        'font_size': 14,
        'font_name': 'Times New Roman',
        'font_bold': False,
        'alignment_vertical': 'center',
        'alignment_horizontal': 'center',
        'alignment_wrap_text': True
    }

    style_config_data = {
        'name': 'data',
        'font_size': 14,
        'font_name': 'Times New Roman',
        'font_italic': False,
        'alignment_vertical': 'center',
        'alignment_horizontal': 'center',
    }

    START_ROW = 1
    START_COL = 1
    START_ROW_DATA = START_ROW + 1
    START_COL_DATA = 1

    def __init__(self, data: list, user):
        self.data = data
        self.user = user
        self.wb = Workbook()
        self._file = io.BytesIO()

    def create_report(self):
        self._set_style_in_wb()
        self.wb.remove(self.wb.active)
        try:
            sheet = self.wb['Объявления']
        except KeyError:
            sheet = self.wb.create_sheet('Объявления')
        self.fill_header(sheet=sheet)
        self.fill_data(sheet=sheet)
        self._save_file()

    def fill_header(self, sheet: Worksheet):
        cur_col = self.START_COL
        cur_row = self.START_ROW
        for header in self.data[0].keys():
            self.set_cell(sheet, header, cur_row, cur_col)
            self.set_width_cols(sheet, cur_col, width=30)
            cur_col += 1
        self.set_style_header(sheet)

    def fill_data(self, sheet: Worksheet):
        cur_col = self.START_COL
        cur_row = self.START_ROW_DATA
        for ad in self.data:
            for ad_value in ad.values():
                self.set_cell(sheet, ad_value, cur_row, cur_col)
                cur_col += 1
            self.set_style_data(sheet, cur_row, cur_col-1)
            cur_col = 1
            cur_row += 1

    # @staticmethod
    # def check_is_link(value):
    #     return re.match('(\w+)://', value)

    def _save_file(self):
        """Сохраняет книгу в файл"""
        self.wb.save(self._file)
        with open(f'{config.MEDIA_ROOT}/{self.get_filename}', 'wb') as file:
            file.write(self._file.getbuffer())

    @property
    def report_path(self):
        """Возвращает файл"""
        return f'{config.MEDIA_ROOT}/{self.get_filename}'

    @property
    def get_filename(self):
        return f'{self.user} объявления Авито за {datetime.date.today()}.xlsx'

    def _set_style_in_wb(self):
        self._get_styles()
        self.wb.add_named_style(self.style_header)
        self.wb.add_named_style(self.style_data)

    def set_cell(self, sheet: Worksheet, value: Any, row: int, column: int) -> Any:
        """Вставляет значение в ячейку"""
        cell = sheet.cell(row=row, column=column)
        # if self.check_is_link(value):
        #     cell.hyperlink = value
        cell.value = value
        return value

    def merge_rows(self, sheet: Worksheet, row: int, row_end: int, col: int) -> None:
        """Объединяет строка в колонке"""
        range_rows = f'{self.get_letter_index(row, col)}:{self.get_letter_index(row_end, col)}'
        sheet.merge_cells(range_string=range_rows)

    def merge_cols(self, sheet: Worksheet, row: int, col: int, col_end: int) -> None:
        """Объединяет колонки в строке"""
        range_cell = f'{self.get_letter_index(row, col)}:{self.get_letter_index(row, col_end)}'
        sheet.merge_cells(range_string=range_cell)

    def get_letter_index(self, row: int, column: int) -> str:
        """Получает полный адрес ячейки в виде строки с литерой"""
        column_letter = get_column_letter(column)
        return f'{column_letter}{row}'

    def set_width_cols(self, sheet: Worksheet, min_col: int, max_col: int = None, width=20):
        """Задает ширину колонки или диапазона колонок"""
        max_col = max_col if max_col else min_col
        for col in range(min_col, max_col + 1):
            sheet.column_dimensions[get_column_letter(col)].width = width

    def set_height_rows(self, sheet: Worksheet, min_row: int, max_row: int, height: int) -> None:
        """Задает высоту строки или диапазона строк"""
        max_row = max_row if max_row else min_row
        for row in range(min_row, max_row + 1):
            sheet.row_dimensions[row].height = height

    # def set_style(self, sheet: Worksheet, footer_row):
    #     """Задает стили документа"""
    #     self.set_style_title_row(sheet)
    #     self.set_style_header(sheet)
    #     self.set_style_data(sheet, footer_row)

    def _get_styles(self):
        self.style_header = self.get_style_header()
        self.style_data = self.get_style_data()

    def get_style_header(self) -> NamedStyle:
        """Возвращает стиль форматирования для заголовков колонок"""
        return get_style(**self.style_config_header)

    def get_style_data(self) -> NamedStyle:
        """Возвращает стиль форматирования для данных таблицы"""
        return get_style(**self.style_config_data)

    def style_cells_as(self,
                       sheet: Worksheet,
                       style: str,
                       min_row: int,
                       min_col: int,
                       max_row: int,
                       max_col: int):
        """Задает стиль диапазону ячеек"""
        rows = sheet.iter_rows(
            min_row=min_row,
            min_col=min_col,
            max_row=max_row,
            max_col=max_col
        )
        for row in rows:
            for cell in row:
                cell.style = style

    def set_style_header(self, sheet: Worksheet):
        row = self.START_ROW
        self.style_cells_as(
            sheet,
            self.style_config_header['name'],
            row,
            self.START_COL,
            row,
            self._get_max_column_header(sheet)
        )

    def _get_max_column_header(self, sheet: Worksheet, min_row=1, max_row=3, max_col=20):
        rows = sheet.iter_rows(
            min_row=min_row,
            min_col=self.START_COL,
            max_row=min_row + 2,
            max_col=max_col
        )
        column = 1
        for row in rows:
            for i, cell in enumerate(row):
                if cell.value:
                    column = i
        return column + 1

    def set_style_data(self, sheet: Worksheet, footer_row: int, max_col):
        self.style_cells_as(
            sheet,
            self.style_config_data['name'],
            self.START_ROW_DATA,
            self.START_COL,
            footer_row,
            max_col
        )
